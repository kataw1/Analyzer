###---Imports-for-views---#

# Additional note we use openpyxl , xlswriter too

#normal django import to return
from django.shortcuts import get_object_or_404, render , redirect
from django.http import HttpResponse, HttpResponseServerError, JsonResponse
from django .contrib.auth.decorators import login_required
#For Ai , and graph and images
from sklearn.linear_model import LinearRegression
from sklearn.linear_model import LogisticRegression , ARDRegression
from sklearn.model_selection import train_test_split
import numpy as np
import matplotlib.pyplot as plt
import matplotlib
from sklearn.tree import DecisionTreeRegressor

from Anlayzer import settings
matplotlib.use('Agg')
import base64
import io
import pandas as pd
import openpyxl
import statistics
import seaborn as sns
from analytic.models import Contact, Favorite, Service , Premium_Service

#imports For regestration system
from django.contrib.auth.models import User ,auth
from django.contrib import messages


#---Functions-Views---#


#______________________Normal Pages URLS______________________#


def home(request): # just to return home page
    is_premium = request.user.groups.filter(name='Premium').exists()
    return render(request,'home_page.html',{'user': request.user,'is_premium': is_premium})


def contact(request): # just to Contact page
    if request.method == 'POST':
        email = request.POST['email']
        message = request.POST['message']
        contact_entry = Contact(email=email, message=message)
        contact_entry.save()
        messages.info(request,'Thank you For contact us')
        return redirect('/contact')
    else: 
        return render(request,'contactUs.html')


@login_required
def services(request):
    # Query all services
    services = Service.objects.all()
    is_premium = request.user.groups.filter(name='Premium').exists()

    # Check favorites for the current user and annotate the services queryset
    if request.user.is_authenticated:
        favorite_services = Favorite.objects.filter(user=request.user, service__in=services)
        favorite_service_ids = favorite_services.values_list('service_id', flat=True)
        for service in services:
            service.is_favorite = service.id in favorite_service_ids
    else:
        messages.info(request,'You should login or sign up')
        return redirect('/login')
    return render(request, 'services.html', {'services': services,'is_premium': is_premium})


def aboutus(request):
    return render(request,'aboutUs.html')

@login_required
def payment(request):
    is_premium = request.user.groups.filter(name='Premium').exists()
    return render(request,'Payment.html',{'is_premium': is_premium})



from django.db.models import Exists, OuterRef

def premium_services(request):
    # Check if the user is authenticated
    if not request.user.is_authenticated:
        messages.info(request, 'You should log in or sign up')
        return redirect('/login')
    
    # Create a subquery to check if the service is a favorite for the current user
    favorite_subquery = Favorite.objects.filter(
        user=request.user,
        service=OuterRef('pk')
    )
    
    # Query all premium services and annotate them with 'is_favorite'
    premium_services = Premium_Service.objects.annotate(
        is_favorite=Exists(favorite_subquery)
    )
    
    return render(request, 'premium_services.html', {'premium_services': premium_services})


#-----------------------------------User Functions-----------------------------------------------#

@login_required
def add_to_fav(request, service_id):
    service = get_object_or_404(Service, pk=service_id)
    fav, created = Favorite.objects.get_or_create(user=request.user, service=service)
    if created:
        messages.info(request, 'Your item has been added successfully!')
    else:
        messages.info(request, 'This item is already in your favorites!')
    return redirect('services')


@login_required
def remove_from_fav(request, service_id):
    service = get_object_or_404(Service, pk=service_id)
    try:
        favorite = Favorite.objects.get(user=request.user, service=service)
        favorite.delete()
        messages.success(request, 'This item has been removed from your favorites.')
    except Favorite.DoesNotExist:
        messages.error(request, 'This item was not in your favorites.')

    # Determine the referer URL to redirect back to the correct page
    referer = request.META.get('HTTP_REFERER')
    if referer and 'view-fav' in referer:
        return redirect('view-fav')
    else:
        return redirect('services')


@login_required
def search_service(request):
    query = request.GET.get('q', '').strip()  # Get and clean the search query from query parameters
    
    if query:
        services = Service.objects.filter(title__icontains=query)  # Search by title (case-insensitive)
    else:
        services = Service.objects.all()  # Display all services if no search query is provided
    
    return render(request, 'services.html', {'services': services, 'query': query})


@login_required
def view_fav(request):
    fav_list = Favorite.objects.filter(user=request.user).select_related('service')
    premium_services = Premium_Service.objects.all()
    
    return render(request, 'view_fav.html', {
        'fav_list': fav_list,
        'premium_services': premium_services
    })



@login_required
def service_filter(request):
    services = Service.objects.all()
    category = request.GET.get('category', None)

    if category:
        services = services.filter(category=category)
    # If no category is specified, do not apply any filtering and retrieve all services

    context = {
        'services': services
    }

    return render(request, 'services.html', context)






#------------------------------------------Regestration Sestym--------------------------------------------s--#

def login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = auth.authenticate(username=username,password=password) #using the authinticate sestym
        if user is not None:
            auth.login(request,user)
            return redirect('/services')
        else:
            messages.info(request,'There is somthing wrong check username or password')
            return redirect('/login')
    return render(request,'login.html')



from django.template.loader import render_to_string
from django.core.mail import EmailMessage
from .models import EmailVerification #important for Verfication

def register(request):
    if request.method == 'POST':
        username = request.POST['username']
        email = request.POST['email']
        password = request.POST['password']
        password2 = request.POST['password2']

        if password == password2:
            if User.objects.filter(email=email).exists():
                messages.info(request, "Email is already in use.")
                return redirect('register')
            elif User.objects.filter(username=username).exists():
                messages.info(request, "Username is already in use.")
                return redirect('register')
            else:
                # Create user instance but do not save yet
                user = User(username=username, email=email)
                user.set_password(password)
                user.is_active = False  # Set user inactive until email confirmation
                user.save()

                # Create email verification instance
                email_verification = EmailVerification.objects.create(user=user)

                # Send verification email
                current_site = request.get_host()
                mail_subject = 'Activate your account'
                message = render_to_string('account_activation_email.html', {
                    'user': user,
                    'domain': current_site,
                    'token': email_verification.token,
                })
                to_email = email
                email = EmailMessage(mail_subject, message, to=[to_email])
                email.send()

                messages.success(request, "Please check your email to activate your account.")
                return redirect('login')
        else:
            messages.info(request, "Passwords do not match.")
            return redirect('register')
    return render(request, 'register.html')



 #This Code We built for Varicfication

def activate(request, token):
    try:
        email_verification = EmailVerification.objects.get(token=token)
        user = email_verification.user
        user.is_active = True
        user.save()
        email_verification.delete()  # Delete the token after activation
        messages.success(request, "Your account has been activated. You can now log in.")
    except EmailVerification.DoesNotExist:
        messages.error(request, "Invalid activation link.")
    
    return redirect('login')




def logout(request):
    auth.logout(request) #just simple request to logout]
    return redirect('/home')

def service_detail(request, service_id):
    # Fetch the service object based on service_id
    service = get_object_or_404(Service, pk=service_id)
    
    # Now you have the 'service' object, you can do further processing or pass it to a template
    return render(request, 'service_detail.html', {'service': service})



#-------------------------------------------------Ecxel_Ai_Services--------------------------------------------------------#

@login_required

def predict_success(request):
    services = Service.objects.all()[:1]
    if request.method == "POST":

        weekly_hours_test = int(request.POST['weekly_hours'])
        difficulty_scale_test = int(request.POST.get("difficulty_scale"))
        self_rating_test = int(request.POST.get("self_rating"))

        weekly_hours = [10, 15, 20, 25, 30]
        difficulty_scale = [3, 6, 7, 8, 9]
        self_rating = [7, 6, 8, 9, 7]
        success_percentage = [65, 50, 80, 90, 75]

        X = [[weekly_hours[i], difficulty_scale[i], self_rating[i]] for i in range(len(weekly_hours))]
        y = success_percentage

        model = LinearRegression()
        model.fit(X, y)

        predicted_success_percentage = model.predict([[weekly_hours_test, difficulty_scale_test, self_rating_test]])

        # Clear the current plot
        plt.clf()

        # Create a DataFrame from the data
        data = pd.DataFrame({'Weekly Hours': weekly_hours, 'Success Percentage': success_percentage})

        # Generate the plot using seaborn
        sns.regplot(x='Weekly Hours', y='Success Percentage', data=data, ci=None)

        # Convert the plot to a base64 string
        img_data = io.BytesIO()
        plt.savefig(img_data, format='png')
        img_data.seek(0)
        base64_img = base64.b64encode(img_data.getvalue()).decode()

        # Return the result to the HTML page
        return render(request, "form.html", {
            "predicted_success_percentage": predicted_success_percentage[0],
            "graph_image": base64_img,
            'services': services
        })

    return render(request, "form.html",{'services': services})


@login_required
def sort_excel_column(request):
    if request.method == 'POST' and 'file' in request.FILES:
        # Read the uploaded Excel file
        file = request.FILES['file']
        try:
            df = pd.read_excel(file)
        except Exception as e:
            return HttpResponse("Error reading Excel file: " + str(e))
        
        # Get the column names to be sorted
        column_names_str = request.POST.get('column_names')
        column_names = [col.strip() for col in column_names_str.split(",")]
        
        # Check if all specified columns exist in the DataFrame
        invalid_columns = [col for col in column_names if col not in df.columns]
        if invalid_columns:
            messages.error(request, f'Invalid column(s) specified: {", ".join(invalid_columns)}')
            return render(request, 'sort.html')
        
        # Get the sort order from the form
        sort_order = request.POST.get('sort_order')
        
        # Sort the dataframe based on each column name and sort order
        sorted_df = df.sort_values(by=column_names, ascending=(sort_order == 'in_order'))
        
        # Convert the sorted dataframe back to Excel file
        output = io.BytesIO()
        writer = pd.ExcelWriter(output, engine='xlsxwriter')
        sorted_df.to_excel(writer, index=False, sheet_name='Sheet1')
        writer.close()
        output.seek(0)
        
        # Set response headers to return the sorted Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=sorted_file.xlsx'
        response.write(output.getvalue())
        
        return response

    # Render the form to upload the file and specify the column names
    return render(request, 'sort.html')



@login_required
def mean_mode_median(request):
    if request.method == "POST":
        if 'file' not in request.FILES:
            messages.error(request, 'No file uploaded.')
            return render(request, 'mean.html')
        
        file = request.FILES['file']
        
        if not file.name.endswith('.xlsx'):
            messages.error(request, 'Uploaded file is not an Excel file (.xlsx).')
            return render(request, 'mean.html')
        
        try:
            df = pd.read_excel(file)
        except Exception as e:
            messages.error(request, f'Error loading Excel file: {str(e)}')
            return render(request, 'mean.html')

        calculations = request.POST.getlist('calculations')
        column = request.POST.get('column')

        # Validate column name
        if column not in df.columns:
            messages.error(request, f'Column "{column}" not found in the uploaded Excel file.')
            return render(request, 'mean.html')

        result = {}
        graph_image = None
        
        if calculations and column:
            if 'mean' in calculations:
                result['mean'] = df[column].mean()
            
            if 'median' in calculations:
                result['median'] = df[column].median()
            
            if 'mode' in calculations:
                result['mode'] = df[column].mode().tolist()
            
            if 'zscore' in calculations:
                result['zscore'] = (df[column] - df[column].mean()) / df[column].std()
            
            if 'std' in calculations:
                result['std'] = df[column].std()
        
        if 'include_graph' in request.POST:
            try:
                plt.cla()
                
                if 'zscore' in result:
                    plt.hist(result['zscore'], bins=20, edgecolor='black')
                    plt.xlabel('Z-score')
                    plt.ylabel('Frequency')
                    plt.title('Distribution of Z-scores')

                elif 'mode' in result:
                    modes = result['mode']
                    mode_counts = pd.Series(modes).value_counts()
                    plt.bar(mode_counts.index, mode_counts.values)
                    plt.xlabel('Mode Value')
                    plt.ylabel('Frequency')
                    plt.title('Mode Frequencies')

                else:
                    plt.bar(result.keys(), result.values())
                    plt.xlabel('Statistic')
                    plt.ylabel('Value')
                    plt.title('Mean, Median, Mode, Standard Deviation, and Z-Score')
                
                img_data = io.BytesIO()
                plt.savefig(img_data, format='png')
                img_data.seek(0)
                base64_img = base64.b64encode(img_data.getvalue()).decode()
                
                graph_image = base64_img
            except Exception as e:
                messages.error(request, f'Error creating graph: {str(e)}')

        return render(request, 'mean.html', {'result': result, 'graph_image': graph_image})
    
    return render(request, 'mean.html')



@login_required
def concatenate(request):
    if request.method == "POST":
        # Check if file is uploaded
        if 'file' not in request.FILES:
            messages.error(request, 'No file uploaded.')
            return render(request, 'concatenate.html')
        
        file = request.FILES['file']
        
        # Check if the uploaded file is an Excel file
        if not file.name.endswith('.xlsx'):
            messages.error(request, 'Uploaded file is not an Excel file (.xlsx).')
            return render(request, 'concatenate.html')
        
        try:
            df = pd.read_excel(file)
        except Exception as e:
            messages.error(request, 'Error: Unable to read the file. Please make sure it is a valid Excel file.')
            return render(request, 'concatenate.html')

        column1 = request.POST.get("column1")
        column2 = request.POST.get("column2")
        
        # Check if columns exist in the DataFrame
        if column1 not in df.columns or column2 not in df.columns:
            messages.error(request, 'Error: One or both of the specified columns do not exist in the Excel file.')
            return render(request, 'concatenate.html')
        
        # Handle concatenation or summation based on column data types
        try:
            # Check if both columns are numeric
            if pd.api.types.is_numeric_dtype(df[column1]) and pd.api.types.is_numeric_dtype(df[column2]):
                df["concatenated_column"] = df[column1].astype(str) + df[column2].astype(str)
            else:
                # If any column is not numeric, compute the sum
                df["summed_column"] = df[column1].fillna(0) + df[column2].fillna(0)
        except Exception as e:
            messages.error(request, 'Error: Unable to process the columns. Please check the column names or the type of data.')
            return render(request, 'concatenate.html')
        
        # Saving the modified DataFrame back to a new Excel file
        output_file = io.BytesIO()
        df.to_excel(output_file, index=False)
        output_file.seek(0)
        
        # Create a response with the old and new data
        response = HttpResponse(output_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=processed_data.xlsx'
        
        return response
    
    return render(request, 'concatenate.html')




@login_required
def remove_duplicates(request):
    if request.method == 'POST':
        # Check if file is uploaded
        if 'file' not in request.FILES:
            messages.error(request, 'No file uploaded.')
            return render(request, 'remove_duplicates.html')
        
        file = request.FILES['file']
        
        # Check if the uploaded file is an Excel file
        if not file.name.endswith('.xlsx'):
            messages.error(request, 'Uploaded file is not an Excel file (.xlsx).')
            return render(request, 'remove_duplicates.html')
        
        try:
            # Read the Excel file into a pandas DataFrame
            df = pd.read_excel(file)
        except Exception as e:
            messages.error(request, f'Error reading Excel file: {str(e)}')
            return render(request, 'remove_duplicates.html')

        # Get the list of column names from user input
        column_names_input = request.POST.get('column_names', '')
        column_names = [col.strip() for col in column_names_input.split(",")]
        
        # Check if specified columns exist in the DataFrame
        invalid_columns = [col for col in column_names if col not in df.columns]
        if invalid_columns:
            messages.error(request, f'Invalid column(s) specified: {", ".join(invalid_columns)}')
            return render(request, 'remove_duplicates.html')
        
        # Remove duplicate values only from the specified columns
        try:
            df[column_names] = df[column_names].apply(lambda x: x.drop_duplicates())
        except Exception as e:
            messages.error(request, f'Error removing duplicates: {str(e)}')
            return render(request, 'remove_duplicates.html')

        # Output the cleaned DataFrame to a new Excel file
        output_file = io.BytesIO()
        df.to_excel(output_file, index=False)
        output_file.seek(0)
        
        # Create a response with the old and new data
        response = HttpResponse(output_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="cleaned_file.xlsx"'
        
        return response

    return render(request, 'remove_duplicates.html')




@login_required
def remove_duplicates_rows(request):
    if request.method == 'POST' and 'file' in request.FILES:
        file = request.FILES['file']
        
        # Check if the uploaded file is an Excel file
        if not file.name.endswith('.xlsx'):
            messages.error(request, 'Uploaded file is not an Excel file (.xlsx).')
            return render(request, 'remove_duplicates_rows.html')

        try:
            # Read the Excel file into a pandas DataFrame
            df = pd.read_excel(file)
        except Exception as e:
            messages.error(request, f'Error reading Excel file: {str(e)}')
            return render(request, 'remove_duplicates_rows.html')

        # Track the initial number of rows for comparison
        initial_rows = df.shape[0]

        # Remove duplicated rows based on all columns
        df.drop_duplicates(inplace=True)

        # Calculate removed rows count
        removed_rows = initial_rows - df.shape[0]

        # Provide feedback on number of rows removed
        if removed_rows > 0:
            messages.success(request, f'Removed {removed_rows} duplicate row(s).')
        else:
            messages.info(request, 'No duplicate rows found.')

        # Output the modified DataFrame to a new Excel file
        output_file = io.BytesIO()
        df.to_excel(output_file, index=False)
        output_file.seek(0)

        # Prepare the response with the modified Excel file
        response = HttpResponse(output_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="modified_file.xlsx"'
        
        return response

    return render(request, 'remove_duplicates_rows.html')


@login_required
def predict_profit(request):
    if request.method == "POST":
        price_test = float(request.POST['price'])
        last_year_sales_test = float(request.POST['last_year_sales'])
        marketing_cost_test = float(request.POST.get('marketing_cost', 0))
        units_sold_test = float(request.POST.get('units_sold', 0))
        competitor_pricing_test = float(request.POST.get('competitor_pricing', 0))
        marketing_rate_test = float(request.POST['marketing_rate'])

        price = [100, 200, 300, 400, 500]
        last_year_sales = [500, 600, 700, 800, 900]
        marketing_cost = [1000, 1200, 1500, 1800, 2000]
        units_sold = [50, 60, 70, 80, 90]
        competitor_pricing = [90, 80, 70, 60, 50]
        marketing_rate = [5, 6, 7, 8, 9]
        profit = [1000, 1500, 1800, 2000, 2200]

        X = np.array([price, last_year_sales, marketing_cost, units_sold, competitor_pricing, marketing_rate]).T
        y = profit

        model = DecisionTreeRegressor(max_depth=5)
        model.fit(X, y)

        predicted_profit = model.predict([[price_test, last_year_sales_test, marketing_cost_test, units_sold_test, competitor_pricing_test, marketing_rate_test]])

        plt.clf()

        fig, axs = plt.subplots(1, 2, figsize=(12, 6))

        # Hexbin plot
        hb = axs[0].hexbin(price, profit, gridsize=20, cmap='inferno')
        axs[0].set_xlabel('Price')
        axs[0].set_ylabel('Profit')
        axs[0].set_title('Profit vs Price')
        cb = fig.colorbar(hb, ax=axs[0])
        cb.set_label('Frequency')

        # Plot the prediction line along with original data points
        x_plot = np.linspace(0, 600, 100)
        y_plot = model.predict(np.column_stack([x_plot, np.repeat(last_year_sales_test, 100), np.repeat(marketing_cost_test, 100), np.repeat(units_sold_test, 100), np.repeat(competitor_pricing_test, 100), np.repeat(marketing_rate_test, 100)]))
        axs[1].plot(x_plot, y_plot, color='r', label='Prediction')
        axs[1].scatter(price, profit, color='b', label='Data points')
        axs[1].set_xlabel('Price')
        axs[1].set_ylabel('Profit')
        axs[1].set_title('Profit Prediction')
        axs[1].legend()

        # Save the figure to a BytesIO object
        img_data = io.BytesIO()
        plt.savefig(img_data, format='png')
        img_data.seek(0)
        base64_img = base64.b64encode(img_data.getvalue()).decode()

        # Return the result to the HTML page
        return render(request, "predict_profit.html", {
            "predicted_profit": predicted_profit[0],
            "graph_image": base64_img
        })

    return render(request, "predict_profit.html")





@login_required
def fill_empty_data(request):
    if request.method == 'POST' and request.FILES.get('file'):
        # Handle file upload
        uploaded_file = request.FILES['file']
        column_names_input = request.POST.get('column_names', '').strip()
        column_names = [col.strip() for col in column_names_input.split(',') if col.strip()]  # Split and clean column names
        fill_value = request.POST.get('fill_value', '').strip()
        
        # Process the uploaded Excel file
        try:
            df = pd.read_excel(uploaded_file, engine='openpyxl')
        except Exception as e:
            messages.error(request, f"Error processing Excel file: {str(e)}")
            return render(request, 'empty_fill.html')

        # Validate column names
        invalid_columns = [col for col in column_names if col not in df.columns]
        if invalid_columns:
            messages.error(request, f"Invalid column(name): {', '.join(invalid_columns)}")
            return render(request, 'empty_fill.html')
        
        # Fill specified columns with the fill value
        for column in column_names:
            df[column] = df[column].fillna(fill_value)
        
        # Create an in-memory Excel file
        output = io.BytesIO()
        writer = pd.ExcelWriter(output, engine='xlsxwriter')
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        writer.close()  # Finalize the XlsxWriter file writing
        
        # Seek to the beginning of the BytesIO buffer
        output.seek(0)
        
        # Set response headers to return the processed Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=processed_file.xlsx'
        response.write(output.getvalue())
        
        # Inform the user that the file has been processed successfully
        messages.info(request, 'File processed successfully and downloaded.')
        
        return response
    
    # Render the form to upload the file and specify the column names
    return render(request, 'empty_fill.html')


from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from PyPDF2 import PdfReader
from summarizer import Summarizer

def summarize_pdf(request):
    if request.method == 'POST' and request.FILES.get('pdf_file'):
        pdf_file = request.FILES['pdf_file']

        # Check if the uploaded file is a PDF
        if not pdf_file.name.endswith('.pdf'):
            error_message = "Only PDF files are allowed."
            return render(request, 'summarize_pdf.html', {'error_message': error_message})

        try:
            # Save the PDF file to the server
            file_name = default_storage.save(pdf_file.name, ContentFile(pdf_file.read()))

            # Extract text content from the PDF file
            with default_storage.open(file_name, 'rb') as f:
                pdf_reader = PdfReader(f)
                text_content = ''
                num_pages = len(pdf_reader.pages)  # Use len(reader.pages) instead of reader.numPages
                for page_num in range(num_pages):
                    page = pdf_reader.pages[page_num]
                    text_content += page.extract_text()

            # Summarize the text content using Summarizer
            model = Summarizer()
            summarized_text = model(text_content)

            # Delete the uploaded PDF file from the server
            default_storage.delete(file_name)

            return render(request, 'summarize_pdf.html', {'summarized_text': summarized_text})

        except PdfReader as e:
            error_message = f"Error reading PDF file: {e}"
            return render(request, 'summarize_pdf.html', {'error_message': error_message})

        except Exception as e:
            error_message = f"Error processing PDF file: {e}"
            return render(request, 'summarize_pdf.html', {'error_message': error_message})

    return render(request, 'summarize_pdf.html')



#THIS FUNCTION NOT GOOD YET #########################################################3



def predict_from_excel(request):
    if request.method == 'POST':
        # Get the uploaded file from the request
        excel_file = request.FILES.get('excel_file')  # Use .get() to avoid MultiValueDictKeyError

        if excel_file:
            # Read the Excel file into a DataFrame
            df = pd.read_excel(excel_file)

            # Get the column names from the request
            column_names = [col.strip() for col in request.POST.get('column_names', '').split(',')]

            # Select only the specified columns
            df = df[column_names]

            # Check if the DataFrame contains the required columns
            if len(column_names) == 0:
                return render(request, 'upload_excel.html', {'error_message': 'Please provide at least one column name'})

            # Check for missing values in the DataFrame
            if df.isnull().values.any():
                return render(request, 'upload_excel.html', {'error_message': 'DataFrame contains missing values'})

            # Prepare input features (X) and target variable (y)
            if len(column_names) == 1:
                X = df.iloc[:, 0].values.reshape(-1, 1)  # Reshape to ensure it's a 2D array
                y = None
            else:
                X = df.iloc[:, :-1]
                y = df.iloc[:, -1]

                # Debugging: Print the shapes of X and y
                print("Shape of X:", X.shape)
                print("Shape of y:", y.shape)

                # Ensure that X and y have the same number of observations
                if X.shape[0] != y.shape[0]:
                    return render(request, 'upload_excel.html', {'error_message': 'Number of observations in X and y must be the same'})

            # Perform linear regression if there's a target variable (y)
            if y is not None:
                # Perform linear regression
                model = LinearRegression()
                model.fit(X, y)

                # Make predictions
                predicted_values = model.predict(X)

                # Plot the data and the regression line
                plt.figure(figsize=(8, 6))
                sns.scatterplot(x=X.columns[0], y=y.name, data=df, color='blue')
                plt.plot(X[X.columns[0]], predicted_values, color='red')
                plt.xlabel(X.columns[0])
                plt.ylabel(y.name)
                plt.title('Linear Regression Prediction')
                
                # Convert the plot to a base64 string
                img_data = io.BytesIO()
                plt.savefig(img_data, format='png')
                img_data.seek(0)
                base64_img = base64.b64encode(img_data.getvalue()).decode()

                # Return the result to the HTML page
                return render(request, 'upload_excel.html', {'predicted_values': predicted_values.tolist(), 'graph_image': base64_img})
            else:
                # No target variable (y), just return the column values
                predicted_values = df.iloc[:, 0].tolist()
                return render(request, 'upload_excel.html', {'predicted_values': predicted_values})

    # Render the upload form if not a POST request or if the file is not provided
    return render(request, 'upload_excel.html')







from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from django.contrib import messages




def color_duplicates(request):
    if request.method == 'POST':
        # Check if file is uploaded
        if 'file' not in request.FILES:
            messages.error(request, 'No file uploaded.')
            return render(request, 'color_duplicates.html')

        file = request.FILES['file']

        # Check if the uploaded file is an Excel file
        if not file.name.endswith('.xlsx'):
            messages.error(request, 'Uploaded file is not an Excel file (.xlsx).')
            return render(request, 'color_duplicates.html')

        try:
            # Read the Excel file into a pandas DataFrame
            df = pd.read_excel(file)
        except Exception as e:
            messages.error(request, f'Error reading Excel file: {str(e)}')
            return render(request, 'color_duplicates.html')

        # Get the column name from user input
        column_name = request.POST.get('column_name', '').strip()

        # Check if the specified column exists in the DataFrame
        if column_name not in df.columns:
            messages.error(request, f'Column "{column_name}" not found in the file.')
            return render(request, 'color_duplicates.html')

        # Identify duplicates in the specified column
        duplicates = df[df.duplicated(subset=[column_name], keep=False)]

        # Create a new Excel file with coloring
        wb = Workbook()
        ws = wb.active

        # Define a red fill
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # Add DataFrame to the new worksheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Check if the current cell's column matches the specified column and its value is a duplicate
                if c_idx == df.columns.get_loc(column_name) + 1:  # +1 because openpyxl is 1-based index
                    if value in duplicates[column_name].values:
                        cell.fill = red_fill

        # Save the workbook to a BytesIO object
        output_file = io.BytesIO()
        wb.save(output_file)
        output_file.seek(0)

        # Create a response with the new Excel file
        response = HttpResponse(output_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="colored_duplicates.xlsx"'

        return response

    return render(request, 'color_duplicates.html')